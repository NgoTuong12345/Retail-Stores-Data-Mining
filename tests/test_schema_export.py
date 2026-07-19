"""Tests for the standardized cross-brand output schema exporter.

Maps internal pipeline field names (Id, Enterprise_Code, Enterprise_Gdt_Code, ...)
to the canonical analyst-facing schema requested for BHX (and future brands),
without touching the internal Id/Name/Name_F fields the engine's own
dedupe/filter invariants (AGENTS.md) rely on.
"""
import csv
import json

import pytest
from openpyxl import load_workbook

from dkkd.schema_export import (
    STANDARD_SCHEMA_FIELDS,
    classify_brand_confidence,
    classify_store_type,
    classify_store_name_pattern,
    compute_duplication_info,
    build_standard_schema,
    sort_chronologically,
    export_standard_schema,
)
from dkkd.config import BrandConfig


def _duplication_status(stores):
    return {rid: info['status'] for rid, info in compute_duplication_info(stores).items()}


# ── classify_store_type ──

_DEFAULT_TYPE_PATTERNS = [
    ('Retail', r'C[ỦƯỬU]A H[ÀA]NG'),
    ('Warehouse', r'\bKHO\b'),
    ('Online', r'\bONLINE\b'),
    ('Services', r'\bFARM\b|NÔNG SẢN'),
]


class TestClassifyStoreType:
    def test_retail_store_name(self):
        name = 'ĐỊA ĐIỂM KINH DOANH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG BÁCH HÓA XANH SỐ 152'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Retail'

    def test_retail_store_name_tolerates_diacritic_typo(self):
        # "CỦA HÀNG" (ủ) is a common typo for "CỬA HÀNG" (ử) seen in real DKKD filings
        name = 'CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỦA HÀNG BÁCH HÓA XANH 204 KÊNH TÂN HÓA'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Retail'

    def test_warehouse_name(self):
        name = 'ĐỊA ĐIỂM KINH DOANH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH – KHO BÁCH HÓA XANH 3240'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Warehouse'

    def test_online_name(self):
        name = 'ĐỊA ĐIỂM KINH DOANH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - BHX ONLINE HCM 5'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Online'

    def test_services_name(self):
        name = 'CHI NHÁNH PHÁT TRIỂN NÔNG SẢN 4K FARM - CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Services'

    def test_bare_entity_with_no_dash_is_corporate(self):
        name = 'CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Corporate'

    def test_unmatched_name_with_dash_is_other(self):
        name = 'CƠ SỞ 1 - CÔNG TY TNHH BHX VIỆT NAM'
        assert classify_store_type(name, _DEFAULT_TYPE_PATTERNS) == 'Other'


# ── classify_brand_confidence ──

_HIGH = [r'C[ỔO] PH[ẦA]N THƯƠNG MẠI\s+B[ÁA]CH H[ÓO]A XANH', r'THƯƠNG MẠI THẾ GIỚI ĐIỆN TỬ']
_LOW = [r'C[ÔO]NG TY TNHH']


class TestClassifyBrandConfidence:
    def test_official_entity_is_high(self):
        name = 'ĐỊA ĐIỂM KINH DOANH CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG SỐ 1'
        assert classify_brand_confidence(name, _HIGH, _LOW) == 'high'

    def test_legacy_mwg_parent_entity_is_high(self):
        name = 'CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI THẾ GIỚI ĐIỆN TỬ - CỬA HÀNG BÁCH HÓA XANH BÌNH LONG 1'
        assert classify_brand_confidence(name, _HIGH, _LOW) == 'high'

    def test_unrelated_tnhh_entity_is_low(self):
        name = 'CÔNG TY TNHH BHX VIỆT NAM'
        assert classify_brand_confidence(name, _HIGH, _LOW) == 'low'

    def test_different_jsc_entity_is_medium(self):
        # Same "BHX" token, different (non-official) joint-stock company — genuinely
        # ambiguous, must not be silently folded into either high or low.
        name = 'CÔNG TY CỔ PHẦN THƯƠNG MẠI VÀ DỊCH VỤ BHX HÀ NỘI'
        assert classify_brand_confidence(name, _HIGH, _LOW) == 'medium'


# ── classify_store_name_pattern ──

from dkkd.schema_export import _DEFAULT_NAME_PATTERN_RULES as _NAME_RULES


class TestClassifyStoreNamePattern:
    def test_province_plus_so_is_province_sequential(self):
        # "...BÁCH HÓA XANH ĐỒNG NAI SỐ 145" — verified Spearman ~0.999 with
        # DKKD registration order within Đồng Nai alone: a per-province counter.
        name = 'ĐỊA ĐIỂM KINH DOANH CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG BÁCH HÓA XANH ĐỒNG NAI SỐ 145'
        assert classify_store_name_pattern(name, 'Retail', _NAME_RULES) == 'Province_Sequential'

    def test_bare_brand_plus_so_is_national_sequential(self):
        # "...CỬA HÀNG BÁCH HÓA XANH SỐ 152" — no province word before SỐ.
        name = 'ĐỊA ĐIỂM KINH DOANH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG BÁCH HÓA XANH SỐ 152'
        assert classify_store_name_pattern(name, 'Retail', _NAME_RULES) == 'National_Sequential'

    def test_province_plus_bare_number_no_so_is_national_sequential(self):
        # "...THANH HOÁ 14515" — province word present but no "SỐ": the running
        # national code just dropped the word once it grew past ~1000.
        name = 'ĐỊA ĐIỂM KINH DOANH CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG BÁCH HÓA XANH THANH HOÁ 14515'
        assert classify_store_name_pattern(name, 'Retail', _NAME_RULES) == 'National_Sequential'

    def test_trailing_house_number_is_street_address(self):
        # "...BÁCH HÓA XANH 41 BÀ ĐIỂM" ends in a ward name, not a digit —
        # the '41' is a house number, not a store sequence code.
        name = 'CHI NHÁNH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG BÁCH HÓA XANH 41 BÀ ĐIỂM'
        assert classify_store_name_pattern(name, 'Retail', _NAME_RULES) == 'Street_Address'

    def test_non_retail_store_type_is_not_applicable(self):
        name = 'ĐỊA ĐIỂM KINH DOANH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH – KHO BÁCH HÓA XANH 3240'
        assert classify_store_name_pattern(name, 'Warehouse', _NAME_RULES) == 'Not_Applicable'


# ── compute_duplication_status ──

class TestComputeDuplicationStatus:
    def test_exact_same_address_is_high(self):
        stores = [
            {'Id': '1', 'Ho_Address': '352 Chu Văn An, Phường 12, Quận Bình Thạnh, TP.HCM'},
            {'Id': '2', 'Ho_Address': '352 Chu Văn An, Phường 12, Quận Bình Thạnh, TP.HCM'},
        ]
        status = _duplication_status(stores)
        assert status['1'] == 'high'
        assert status['2'] == 'high'

    def test_substring_match_is_medium(self):
        stores = [
            {'Id': '1', 'Ho_Address': '123 Nguyễn Văn Cừ, Phường 4, Quận 5, TP.HCM'},
            {'Id': '2', 'Ho_Address': '123 Nguyễn Văn Cừ'},
        ]
        status = _duplication_status(stores)
        assert status['1'] == 'medium'
        assert status['2'] == 'medium'

    def test_unique_address_is_low(self):
        stores = [
            {'Id': '1', 'Ho_Address': '10 Lê Lợi, Quận 1, TP.HCM'},
            {'Id': '2', 'Ho_Address': '99 Trần Hưng Đạo, Quận 5, TP.HCM'},
        ]
        status = _duplication_status(stores)
        assert status['1'] == 'low'
        assert status['2'] == 'low'

    def test_blank_address_is_low_not_grouped_together(self):
        # Multiple blank addresses must not count as a duplicate group.
        stores = [
            {'Id': '1', 'Ho_Address': ''},
            {'Id': '2', 'Ho_Address': ''},
        ]
        status = _duplication_status(stores)
        assert status['1'] == 'low'
        assert status['2'] == 'low'

    def test_no_row_is_dropped(self):
        # The whole point: every input record gets a status, none excluded.
        stores = [
            {'Id': '1', 'Ho_Address': '352 Chu Văn An, TP.HCM'},
            {'Id': '2', 'Ho_Address': '352 Chu Văn An, TP.HCM'},
            {'Id': '3', 'Ho_Address': '99 Trần Hưng Đạo, TP.HCM'},
        ]
        status = _duplication_status(stores)
        assert set(status.keys()) == {'1', '2', '3'}


# ── compute_duplication_info (status + reason) ──

class TestComputeDuplicationInfo:
    def test_high_reason_names_the_matching_id(self):
        stores = [
            {'Id': '5177142', 'Ho_Address': '352 Chu Văn An, Phường 12, Quận Bình Thạnh, TP.HCM'},
            {'Id': '10961092', 'Ho_Address': '352 Chu Văn An, Phường 12, Quận Bình Thạnh, TP.HCM'},
        ]
        info = compute_duplication_info(stores)
        assert info['5177142']['status'] == 'high'
        assert info['5177142']['reason'] == 'Exact address match with Id 10961092'
        assert info['10961092']['reason'] == 'Exact address match with Id 5177142'

    def test_high_reason_lists_all_other_ids_in_a_larger_group(self):
        addr = '35-37 Phạm Hữu Lầu, Khu phố 2, Phường Phú Mỹ, Quận 7, TP.HCM'
        stores = [
            {'Id': 'A', 'Ho_Address': addr},
            {'Id': 'B', 'Ho_Address': addr},
            {'Id': 'C', 'Ho_Address': addr},
        ]
        info = compute_duplication_info(stores)
        assert info['A']['reason'] == 'Exact address match with Id B, C'

    def test_medium_reason_names_the_overlapping_id(self):
        stores = [
            {'Id': '1', 'Ho_Address': '123 Nguyễn Văn Cừ, Phường 4, Quận 5, TP.HCM'},
            {'Id': '2', 'Ho_Address': '123 Nguyễn Văn Cừ'},
        ]
        info = compute_duplication_info(stores)
        assert info['1']['status'] == 'medium'
        assert 'Id 2' in info['1']['reason']
        assert info['2']['status'] == 'medium'
        assert 'Id 1' in info['2']['reason']

    def test_low_reason_is_blank(self):
        stores = [
            {'Id': '1', 'Ho_Address': '10 Lê Lợi, Quận 1, TP.HCM'},
            {'Id': '2', 'Ho_Address': '99 Trần Hưng Đạo, Quận 5, TP.HCM'},
        ]
        info = compute_duplication_info(stores)
        assert info['1'] == {'status': 'low', 'reason': ''}
        assert info['2'] == {'status': 'low', 'reason': ''}

    def test_compute_duplication_status_stays_consistent_with_info(self):
        stores = [
            {'Id': '1', 'Ho_Address': '352 Chu Văn An, TP.HCM'},
            {'Id': '2', 'Ho_Address': '352 Chu Văn An, TP.HCM'},
            {'Id': '3', 'Ho_Address': '99 Trần Hưng Đạo, TP.HCM'},
        ]
        status = _duplication_status(stores)
        info = compute_duplication_info(stores)
        assert status == {rid: v['status'] for rid, v in info.items()}


# ── build_standard_schema ──

_CONFIG = BrandConfig(
    slug='bach-hoa-xanh',
    name='Bách Hóa Xanh',
    brand_regex='BÁCH HÓA XANH|BHX',
    classification={
        'schema_export': {
            'official_entity_patterns': _HIGH,
            'low_confidence_patterns': _LOW,
            'store_type_patterns': [list(p) for p in _DEFAULT_TYPE_PATTERNS],
        }
    },
)

_SAMPLE_RECORD = {
    'Id': '5177142',
    'Enterprise_Code': '0310471746',
    'Enterprise_Gdt_Code': '00152',
    'Name': 'ĐỊA ĐIỂM KINH DOANH CÔNG TY CỔ PHẦN THƯƠNG MẠI BÁCH HÓA XANH - CỬA HÀNG BÁCH HÓA XANH SỐ 152',
    'Establishment_Date': '2018-08-12',
    'Establishment_Year': 2018,
    'Date_Confidence': 'global',
    'City_Id': '79',
    'District_Id': '765',
    'Ward_Id': '27058',
    'Ho_Address': '352 Chu Văn An, Phường 12, Quận Bình Thạnh, Thành phố Hồ Chí Minh, Việt Nam',
    'City_Name': 'Thành phố Hồ Chí Minh',
    'District_Name': 'Quận Bình Thạnh',
    'Ward_Name': 'Phường 12',
    'Legal_First_Name': 'NGUYỄN ĐỨC TÀI',
}


class TestBuildStandardSchema:
    def test_field_set_matches_spec_exactly(self):
        rows = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])
        assert set(rows[0].keys()) == set(STANDARD_SCHEMA_FIELDS)

    def test_identity_fields_mapped(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['DKKD_internal_id'] == '5177142'
        assert row['DKKD_enterprise_id'] == '0310471746'
        assert row['MST_gdt_code'] == '00152'

    def test_brand_name_and_confidence(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['store_brand_name'] == 'Bách Hóa Xanh'
        assert row['store_brand_name_confidence'] == 'high'

    def test_store_type_and_name_details(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['store_type'] == 'Retail'
        assert 'CỬA HÀNG BÁCH HÓA XANH SỐ 152' in row['store_name_details']

    def test_store_name_pattern(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['store_name_pattern'] == 'National_Sequential'

    def test_duplication_status_and_reason_low_for_unique_address(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['duplication_status'] == 'low'
        assert row['duplication_reason'] == ''

    def test_duplication_status_and_reason_high_when_address_shared(self):
        other = dict(_SAMPLE_RECORD, Id='10961092', Name='... CỬA HÀNG BÁCH HÓA XANH 14396')
        rows = build_standard_schema(_CONFIG, [_SAMPLE_RECORD, other])
        assert rows[0]['duplication_status'] == 'high'
        assert rows[0]['duplication_reason'] == 'Exact address match with Id 10961092'
        assert rows[1]['duplication_reason'] == 'Exact address match with Id 5177142'

    def test_date_fields_derived_correctly(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['establishment_date_inferred'] == '2018-08-12'
        assert row['establishment_date_confidence'] == 'global'
        assert row['establishment_month'] == 8
        assert row['establishment_quarter'] == 3
        assert row['establishment_year'] == 2018

    def test_geo_and_legal_fields_passthrough(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['province_city_id'] == '79'
        assert row['district_id'] == '765'
        assert row['ward_id'] == '27058'
        assert row['full_address'] == _SAMPLE_RECORD['Ho_Address']
        assert row['province_city_name'] == 'Thành phố Hồ Chí Minh'
        assert row['district_name'] == 'Quận Bình Thạnh'
        assert row['ward_name'] == 'Phường 12'
        assert row['legal_representative'] == 'NGUYỄN ĐỨC TÀI'

    def test_region_defaults_blank_when_record_lacks_it(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['province_region'] == ''
        assert row['province_region_post_reform'] == ''

    def test_region_passthrough_when_record_has_it(self):
        record = dict(_SAMPLE_RECORD, Region='Đông Nam Bộ', Region_Post_Reform='Đông Nam Bộ')
        row = build_standard_schema(_CONFIG, [record])[0]
        assert row['province_region'] == 'Đông Nam Bộ'
        assert row['province_region_post_reform'] == 'Đông Nam Bộ'

    def test_region_and_region_post_reform_can_differ(self):
        record = dict(_SAMPLE_RECORD, Region='Tây Bắc Bộ', Region_Post_Reform='Đông Bắc Bộ')
        row = build_standard_schema(_CONFIG, [record])[0]
        assert row['province_region'] == 'Tây Bắc Bộ'
        assert row['province_region_post_reform'] == 'Đông Bắc Bộ'

    def test_missing_establishment_date_leaves_month_quarter_blank(self):
        record = dict(_SAMPLE_RECORD, Establishment_Date=None)
        row = build_standard_schema(_CONFIG, [record])[0]
        assert row['establishment_month'] == ''
        assert row['establishment_quarter'] == ''

    def test_falls_back_to_default_patterns_when_config_has_no_schema_export(self):
        bare_config = BrandConfig(slug='x', name='X', brand_regex='X', classification={})
        rows = build_standard_schema(bare_config, [_SAMPLE_RECORD])
        # No official_entity_patterns configured => can't be 'high', but must not crash
        assert rows[0]['store_brand_name_confidence'] == 'medium'


# ── sort_chronologically ──

class TestSortChronologically:
    def test_orders_oldest_first(self):
        rows = [
            {'DKKD_internal_id': '3', 'establishment_date_inferred': '2020-01-01'},
            {'DKKD_internal_id': '1', 'establishment_date_inferred': '2015-06-15'},
            {'DKKD_internal_id': '2', 'establishment_date_inferred': '2018-03-10'},
        ]
        ordered = sort_chronologically(rows)
        assert [r['DKKD_internal_id'] for r in ordered] == ['1', '2', '3']

    def test_blank_dates_sort_last(self):
        rows = [
            {'DKKD_internal_id': '1', 'establishment_date_inferred': ''},
            {'DKKD_internal_id': '2', 'establishment_date_inferred': '2015-06-15'},
        ]
        ordered = sort_chronologically(rows)
        assert [r['DKKD_internal_id'] for r in ordered] == ['2', '1']

    def test_ties_break_on_internal_id(self):
        rows = [
            {'DKKD_internal_id': '9', 'establishment_date_inferred': '2020-01-01'},
            {'DKKD_internal_id': '2', 'establishment_date_inferred': '2020-01-01'},
        ]
        ordered = sort_chronologically(rows)
        assert [r['DKKD_internal_id'] for r in ordered] == ['2', '9']


# ── export_standard_schema (file I/O) ──

_MINIMAL_CONFIG_YAML = (
    "slug: test-brand\n"
    "name: Test Brand\n"
    "brand_regex: 'TEST'\n"
    "classification:\n"
    "  schema_export:\n"
    "    official_entity_patterns:\n"
    "      - 'OFFICIAL CO'\n"
    "    low_confidence_patterns:\n"
    "      - 'TNHH'\n"
)


def _setup_brand(tmp_path, records):
    bd = tmp_path / 'test-brand'
    bd.mkdir()
    (bd / 'output').mkdir()
    (bd / 'config.yaml').write_text(_MINIMAL_CONFIG_YAML, encoding='utf-8')
    (bd / 'checkpoint.json').write_text(json.dumps(records), encoding='utf-8')
    return bd


class TestExportStandardSchema:
    def test_writes_csv_with_expected_header_and_row_count(self, tmp_path):
        records = [['1', dict(_SAMPLE_RECORD, Id='1', Name='OFFICIAL CO - CỬA HÀNG 1')]]
        _setup_brand(tmp_path, records)

        out_path = export_standard_schema('test-brand', brands_dir=tmp_path)

        assert out_path.exists()
        assert out_path.name == 'test-brand_standard_schema.csv'
        with open(out_path, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            assert reader.fieldnames == STANDARD_SCHEMA_FIELDS
            rows = list(reader)
        assert len(rows) == 1
        assert rows[0]['DKKD_internal_id'] == '1'
        assert rows[0]['store_brand_name_confidence'] == 'high'

    def test_exported_rows_are_chronologically_ordered(self, tmp_path):
        records = [
            ['3', dict(_SAMPLE_RECORD, Id='3', Name='OFFICIAL CO - CỬA HÀNG 3', Establishment_Date='2022-05-01')],
            ['1', dict(_SAMPLE_RECORD, Id='1', Name='OFFICIAL CO - CỬA HÀNG 1', Establishment_Date='2015-01-01')],
            ['2', dict(_SAMPLE_RECORD, Id='2', Name='OFFICIAL CO - CỬA HÀNG 2', Establishment_Date='2018-01-01')],
        ]
        _setup_brand(tmp_path, records)

        csv_path = export_standard_schema('test-brand', brands_dir=tmp_path)
        with open(csv_path, encoding='utf-8-sig') as f:
            ids = [row['DKKD_internal_id'] for row in csv.DictReader(f)]
        assert ids == ['1', '2', '3']

        wb = load_workbook(csv_path.with_suffix('.xlsx'))
        ws = wb.active
        xlsx_ids = [r[0].value for r in ws.iter_rows(min_row=2)]
        assert xlsx_ids == ['1', '2', '3']

    def test_writes_excel_safe_xlsx_sibling_with_leading_zeros_preserved(self, tmp_path):
        record = dict(
            _SAMPLE_RECORD,
            Id='1',
            Name='OFFICIAL CO - CỬA HÀNG 1',
            Enterprise_Code='0037610658',
            Enterprise_Gdt_Code='00152',
        )
        _setup_brand(tmp_path, [['1', record]])

        csv_path = export_standard_schema('test-brand', brands_dir=tmp_path)
        xlsx_path = csv_path.with_suffix('.xlsx')

        assert xlsx_path.exists()
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == STANDARD_SCHEMA_FIELDS

        row = {h: c for h, c in zip(header, next(ws.iter_rows(min_row=2, max_row=2)))}
        assert row['DKKD_internal_id'].value == '1'
        assert row['DKKD_enterprise_id'].value == '0037610658'
        assert row['MST_gdt_code'].value == '00152'
        # Text format ('@') is what stops Excel from re-inferring these as numbers
        assert row['DKKD_internal_id'].number_format == '@'
        assert row['DKKD_enterprise_id'].number_format == '@'
        assert row['MST_gdt_code'].number_format == '@'

    def test_writes_metadata_json_alongside_export(self, tmp_path):
        records = [
            ['1', dict(_SAMPLE_RECORD, Id='1', Name='OFFICIAL CO - CỬA HÀNG 1')],
            ['2', dict(_SAMPLE_RECORD, Id='2', Name='OFFICIAL CO - CỬA HÀNG 2')],
        ]
        bd = _setup_brand(tmp_path, records)

        export_standard_schema('test-brand', brands_dir=tmp_path)

        metadata_path = bd / 'output' / 'metadata.json'
        assert metadata_path.exists()
        metadata = json.loads(metadata_path.read_text(encoding='utf-8'))
        assert metadata['brand_slug'] == 'test-brand'
        assert metadata['brand_name'] == 'Test Brand'
        assert metadata['row_count'] == 2
        assert metadata['source_checkpoint'] == 'checkpoint.json'
        assert metadata['generated_at']


# ── store_role / host_store columns ──

class TestTenantColumns:
    def test_new_fields_in_schema(self):
        assert 'store_role' in STANDARD_SCHEMA_FIELDS
        assert 'host_store' in STANDARD_SCHEMA_FIELDS

    def test_reads_role_and_host_from_record(self):
        rec = dict(_SAMPLE_RECORD, store_role='in_brand_tenant', host_store='Co.opXtra Sư Vạn Hạnh')
        row = build_standard_schema(_CONFIG, [rec])[0]
        assert row['store_role'] == 'in_brand_tenant'
        assert row['host_store'] == 'Co.opXtra Sư Vạn Hạnh'

    def test_defaults_when_record_untagged(self):
        row = build_standard_schema(_CONFIG, [_SAMPLE_RECORD])[0]
        assert row['store_role'] == 'own_store'
        assert row['host_store'] == ''
